#!/usr/bin/env python3
import argparse
import re
import math

from openpyxl import Workbook, load_workbook

"""
Script to extract strings around two given indexes of a sequence.


Copyright 2020 Margherita Maria Ferrari.


This file is part of RegionsExtractor.

RegionsExtractor is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

RegionsExtractor is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with RegionsExtractor.  If not, see <http://www.gnu.org/licenses/>.
"""


class RegionsExtractor:
    @classmethod
    def __key_sort(cls, x):
        # return x[1].get('count', 0)
        return x[1].get('weight', 0)

    @classmethod
    def __get_regions(cls, seq, idx, window_length, max_padding=0):
        if max_padding < 0:
            max_padding = 0

        r1 = list()
        r2 = list()
        start = idx - window_length if idx - window_length >= 0 else 0
        end = idx + window_length if idx + window_length <= len(seq) else len(seq)

        for i in range(0, max_padding + 1):
            if start - i >= 0:
                r1.append(seq[start - i:idx - i])
            if end + i <= len(seq):
                r2.append(seq[idx + i:end + i])

        return r1, r2

    @classmethod
    def __add_region_info(cls, regions, key_name, region, seq, start_idx, end_idx):
        if region not in regions[key_name].keys():
            item = {'count': 1,
                    'gene': len(re.findall(r'(?=' + region + r')', seq[start_idx:end_idx]))
                    }
            for i in ['A', 'C', 'G', 'T']:
                item['count_' + i.lower()] = region.count(i)
            regions[key_name][region] = item
            regions[key_name]['unique_wnd'] = regions[key_name].get('unique_wnd', 0) + 1
        else:
            regions[key_name][region]['count'] = regions[key_name][region].get('count', 0) + 1

        regions[key_name]['total_wnd'] = regions[key_name].get('total_wnd', 0) + 1

        for i in ['A', 'C', 'G', 'T']:
            k = 'count_' + i.lower()
            regions[key_name][k] = regions[key_name].get(k, 0) + regions[key_name][region].get(k, 0)

    @classmethod
    def get_args(cls):
        parser = argparse.ArgumentParser(description='Regions extractor')
        parser.add_argument('-f', '--input-fasta', metavar='FASTA_IN_FILE', type=str, required=True,
                            help='FASTA input file', default=None)
        parser.add_argument('-b', '--input-bed', metavar='BED_IN_FILE', type=str, required=True,
                            help='BED input file', default=None)
        parser.add_argument('-ws', '--window_length_small', metavar='WINDOW_LENGTH_SMALL', type=int, required=False,
                            help='Number of nucleotides in small single region', default=5)
        parser.add_argument('-wl', '--window_length_large', metavar='WINDOW_LENGTH_LARGE', type=int, required=False,
                            help='Number of nucleotides in large single region', default=10)
        parser.add_argument('-o', '--output-file', metavar='OUTPUT_FILE', type=str, required=False,
                            help='Output XLSX file', default='output')
        parser.add_argument('-m', '--merge-regions', required=False, action='store_true',
                            help='Merge first 1st and 2nd regions (resp. 3rd and 4th)')
        parser.add_argument('-s', '--start-index', metavar='START_INDEX', type=int, required=True,
                            help='Start index of gene region', default=0)
        parser.add_argument('-e', '--end-index', metavar='END_INDEX', type=int, required=True,
                            help='End index of gene region', default=0)
        parser.add_argument('-t', '--threshold', type=int, required=False,
                            help='Threshold on word counts', default=20)
        parser.add_argument('-p', '--max-padding', type=int, required=False,
                            help='Maximum number of nucleotides to be padded for each region', default=0)
        parser.add_argument('-l', '--compute_large_region', required=False, action='store_true',
                            help='Compute output file for large region')
        return parser.parse_args()

    @classmethod
    def extract_regions(cls, fasta_in, bed_in, start_idx, end_idx, window_length=5, out_pref='output',
                        num_regions=4, padding=0, bed_extra=False):
        regions = {}
        out_file = out_pref + '_w' + str(window_length) + '_weight.xlsx'
        out_file_extra = out_pref + '_w' + str(window_length) + '_weight_extra.xlsx'

        for i in range(num_regions):
            regions['Region ' + str(i + 1)] = dict()

        with open(fasta_in, 'r') as fin:
            fin.readline()
            seq = fin.readline().strip().upper()

        with open(bed_in, 'r') as fin:
            line = fin.readline()
            rloops_count = 0

            if bed_extra:
                fout = open(bed_in + '_extra.bed', 'w')

            while line:
                parts = line.strip().split('\t')
                idx_1 = int(parts[1])
                idx_2 = int(parts[2])

                if idx_1 > idx_2:
                    raise AssertionError('First index must be less or equal to second index')

                i = 0
                while ((idx_2 - (idx_1 + i)) % window_length) != 0:  # modify first index in R-loop
                    if i >= 0:
                        i += 1

                        if (idx_1 - i) < 0:
                            continue

                    i *= -1

                idx_1 += i

                if bed_extra:
                    parts[1] = str(idx_1)
                    fout.write('\t'.join(parts) + '\n')

                regions_1, regions_2 = cls.__get_regions(seq, idx_1, window_length, padding)
                regions_3, regions_4 = cls.__get_regions(seq, idx_2, window_length, padding)

                if num_regions == 2:
                    cls.__add_region_info(regions, 'Region 1', regions_1[0] + regions_2[0], seq, start_idx, end_idx)
                    cls.__add_region_info(regions, 'Region 2', regions_3[0] + regions_4[0], seq, start_idx, end_idx)
                else:
                    for r in regions_1:
                        cls.__add_region_info(regions, 'Region 1', r, seq, start_idx, end_idx)

                    for r in regions_2:
                        cls.__add_region_info(regions, 'Region 2', r, seq, start_idx, end_idx)

                    for r in regions_3:
                        cls.__add_region_info(regions, 'Region 3', r, seq, start_idx, end_idx)

                    for r in regions_4:
                        cls.__add_region_info(regions, 'Region 4', r, seq, start_idx, end_idx)

                rloops_count += 1
                line = fin.readline()

            if bed_extra:
                fout.close()

        for r in regions.values():
            for k, v in r.items():
                if '_' in k:
                    continue
                if v.get('gene', 0) == 0:
                    v['weight'] = -1  # string not in gene
                else:
                    v['weight'] = v.get('count', 0) / (v.get('gene', 0) * rloops_count)

        wb = Workbook(write_only=True)
        wb_extra = Workbook(write_only=True)
        ws_extra_r2_extra = wb_extra.create_sheet('Region 2 Extra')
        ws_extra_r3_extra = wb_extra.create_sheet('Region 3 Extra')

        for k, v in regions.items():
            add_to_extra = True
            prev_weight = -math.inf
            count_keys = [c for c in v.keys() if 'count' in c]  # count_A, count_C, count_G, count_T for Region 1,2,3,4
            ws = wb.create_sheet(k)
            ws_extra = wb_extra.create_sheet(k)

            header = ['R-LOOPS', 'TOTAL WINDOWS', 'UNIQUE WINDOWS']
            header.extend([i.replace('_', ' ').upper() for i in count_keys])
            ws.append(header)

            header_values = [rloops_count, v.get('total_wnd', 0), v.get('unique_wnd', 0)]
            header_values.extend([v.get(i, 0) for i in count_keys])
            ws.append(header_values)

            ws.append(())

            info_header = ['WINDOW', 'COUNT']
            info_header.extend([i.replace('_', ' ').upper() for i in count_keys])
            info_header.append('GENE ' + str(start_idx) + ' - ' + str(end_idx) + ' NT')
            info_header.append('WEIGHT')
            ws.append(info_header)
            # i is the key in the dict v
            sorted_items = [(i, j) for i, j in v.items() if i not in count_keys and '_' not in i]
            sorted_items.sort(key=cls.__key_sort, reverse=True)
            for wnd, wnd_info in sorted_items:
                info = [wnd, wnd_info.get('count', 0)]
                info.extend([wnd_info.get(i, 0) for i in count_keys])
                info.append(wnd_info.get('gene', 0))
                info.append(wnd_info.get('weight', 0))
                ws.append(info)

                if k.endswith('2'):
                    ws_extra_r2_extra.append(info)
                elif k.endswith('3'):
                    ws_extra_r3_extra.append(info)

                if add_to_extra and \
                        (wnd_info.get('weight', 0) >= 0.01 or abs(prev_weight - wnd_info.get('weight', 0)) <= 0.001):
                    ws_extra.append(info)
                    prev_weight = wnd_info.get('weight', 0)
                else:
                    add_to_extra = False

            ws.close()
            ws_extra.close()

        wb.save(out_file)

        ws_extra_r2_extra.close()
        ws_extra_r3_extra.close()
        wb_extra.move_sheet('Region 2 Extra', 5)
        wb_extra.move_sheet('Region 3 Extra', 6)
        wb_extra.save(out_file_extra)

        return out_file

    @classmethod
    def __get_counts(cls, wb, sub_word):
        counts = [0] * len(wb.sheetnames)

        for n, ws_name in enumerate(wb.sheetnames):
            ws = wb[ws_name]

            for row in ws.iter_rows(min_row=5, max_col=1):
                for cell in row:
                    seq = str(cell.value)
                    counts[n] += int(ws['B' + str(cell.row)].value) * len(re.findall(r'(?=' + sub_word + r')', seq))
                    # counts[n-1] += int(ws['B' + str(cell.row)].value) * word.count(seq)
                    # word.count(seq) counts with multiplicity

        return counts

    @classmethod
    def compare_windows(cls, in_file_small, in_file_large, threshold, window_large):
        wb_small = load_workbook(in_file_small)
        wb_large = load_workbook(in_file_large, read_only=True)

        for ws_name in wb_small.sheetnames:
            ws = wb_small[ws_name]
            ws['I4'] = 'REGION 1 - ' + str(window_large) + 'NT'
            ws['J4'] = 'W_REGION 1 - ' + str(window_large) + 'NT'
            ws['K4'] = 'REGION 2 - ' + str(window_large) + 'NT'
            ws['L4'] = 'W_REGION 2 - ' + str(window_large) + 'NT'
            ws['M4'] = 'REGION 3 - ' + str(window_large) + 'NT'
            ws['N4'] = 'W_REGION 3 - ' + str(window_large) + 'NT'
            ws['O4'] = 'REGION 4 - ' + str(window_large) + 'NT'
            ws['P4'] = 'W_REGION 4 - ' + str(window_large) + 'NT'

            for row in ws.iter_rows(min_row=5, max_col=1):

                for cell in row:
                    word = str(cell.value)

                    if not word or int(ws['B' + str(cell.row)].value) < threshold:
                        continue

                    counts_region = cls.__get_counts(wb_large, word)

                    for i in range(len(counts_region)):
                        if i == 0:
                            cols = ('I', 'J')
                        elif i == 1:
                            cols = ('K', 'L')
                        elif i == 2:
                            cols = ('M', 'N')
                        else:
                            cols = ('O', 'P')

                        ws[cols[0] + str(cell.row)] = counts_region[i]
                        if wb_small[ws_name]['G' + str(cell.row)].value == 0:
                            ws[cols[1] + str(cell.row)] = -1
                        else:
                            ws[cols[1] + str(cell.row)] = (
                                    wb_small[ws_name]['B' + str(cell.row)].value * counts_region[i] /
                                    ((wb_small[ws_name]['A2'].value ** 2) *
                                     wb_small[ws_name]['G' + str(cell.row)].value))

        wb_small.save(in_file_small)


if __name__ == '__main__':
    args = vars(RegionsExtractor.get_args())
    if args.get('merge_regions', False):
        # no padding for merge
        RegionsExtractor.extract_regions(args.get('input_fasta', None), args.get('input_bed', None),
                                         args.get('start_index', 0), args.get('end_index', 0),
                                         args.get('window_length_small', 5), args.get('output_file', 'output'), 2)
    else:
        small_window = RegionsExtractor.extract_regions(args.get('input_fasta', None), args.get('input_bed', None),
                                                        args.get('start_index', 0), args.get('end_index', 0),
                                                        args.get('window_length_small', 5),
                                                        args.get('output_file', 'output'), 4,
                                                        args.get('max_padding', 0), True)

        if args.get('compute_large_window', False):
            large_window = RegionsExtractor.extract_regions(args.get('input_fasta', None), args.get('input_bed', None),
                                                            args.get('start_index', 0), args.get('end_index', 0),
                                                            args.get('window_length_large', 10),
                                                            args.get('output_file', 'output'), 4,
                                                            args.get('max_padding', 0), True)

            RegionsExtractor.compare_windows(small_window, large_window, args.get('threshold', 20),
                                             args.get('window_length_large', 10))
