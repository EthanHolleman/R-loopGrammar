#!/usr/bin/env python3
import argparse
import re
import math

import openpyxl
from openpyxl import Workbook, load_workbook

"""
Script to select most important tuples in each region.


Copyright 2020 Margherita Maria Ferrari.


This file is part of RegionsExtractor.

RegionsExtractor is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

RegionsExtractor is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with RegionsExtractor.  If not, see <http://www.gnu.org/licenses/>.
"""


class RegionsThreshold:
    @classmethod
    def __threshold_weight(cls, xlsx_in, out_file):
        wb = openpyxl.load_workbook(xlsx_in, read_only=True)
        wb_out = Workbook(write_only=True)

        for i, ws_name in enumerate(wb.sheetnames):
            ws_out = wb_out.create_sheet(ws_name)
            prev_weight = -math.inf

            for row in wb.worksheets[i].iter_rows(values_only=True):
                weight = float(row[len(row) - 1])

                if weight >= 0.01 or abs(prev_weight - weight) <= 0.001:
                    ws_out.append(row)
                    prev_weight = weight
                else:
                    ws_out.close()
                    break

        wb_out.save(out_file)

    @classmethod
    def __threshold_shannon(cls, xlsx_in, out_file):
        wb = openpyxl.load_workbook(xlsx_in, read_only=True)
        wb_out = Workbook(write_only=True)

        for i, ws_name in enumerate(wb.sheetnames):
            count = 1
            tmp_sum = 0
            max_weight = 0
            ws_out = wb_out.create_sheet(ws_name)
            prev_entropy = -math.inf

            for row in wb.worksheets[i].iter_rows(values_only=True):
                weight = float(row[len(row) - 1])

                if count == 1 or max_weight == 0:
                    max_weight = weight

                p = weight / max_weight
                tmp = -p * math.log(p, 10)
                tmp_sum += tmp
                entropy = tmp_sum / count

                if entropy >= prev_entropy:
                    new_row = list(row)
                    new_row.append(entropy)
                    ws_out.append(new_row)
                    prev_entropy = entropy
                else:
                    ws_out.close()
                    break

                count += 1

        wb_out.save(out_file)

    @classmethod
    def get_args(cls):
        parser = argparse.ArgumentParser(description='Regions threshold')
        parser.add_argument('-i', '--input-xlsx', metavar='XLSX_IN_FILE', type=str, required=False,
                            help='XLSX input file', default=None)
        parser.add_argument('-s', '--shannon-entropy', required=False, action='store_true',
                            help='Use Shannon entropy')
        parser.add_argument('-o', '--output-file', metavar='OUTPUT_FILE', type=str, required=False,
                            help='Output XLSX file', default='output')
        return parser.parse_args()

    @classmethod
    def extract_regions(cls, xlsx_in, out_file='output.xlsx', shannon_entropy=False):
        if shannon_entropy:
            cls.__threshold_shannon(xlsx_in, out_file)
        else:
            cls.__threshold_weight(xlsx_in, out_file)


if __name__ == '__main__':
    args = vars(RegionsThreshold.get_args())
    RegionsThreshold.extract_regions(args.get('input_xlsx', None), args.get('output_file', 'output.xlsx'),
                                     args.get('shannon_entropy', False))
