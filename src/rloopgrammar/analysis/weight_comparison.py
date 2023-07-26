import argparse
import dataclasses
import math
import openpyxl
import collections

from typing import *

def compare_weighted_regions(weighted_region_xlsx1, weighted_region_xlsx2):
    workbook1 = openpyxl.load_workbook(weighted_region_xlsx1, read_only=True)
    workbook2 = openpyxl.load_workbook(weighted_region_xlsx2, read_only=True)
    
    workbook_reference = collections.defaultdict(lambda: [])
    for i, ws_name in enumerate(workbook1.sheetnames):
        for row in workbook1.worksheets[i].iter_rows(values_only=True):
            workbook_reference[ws_name].append(row)

    for i, ws_name in enumerate(workbook2.sheetnames):
        for j, row in enumerate(workbook2.worksheets[i].iter_rows(values_only=True)):
            if workbook_reference[ws_name][j] != row:
                print(f"{ws_name} {row} != f{workbook_reference[ws_name][j]}")


parser = argparse.ArgumentParser(
                    prog = 'weight_comparison',
                    description = 'Compares two or more weighted regions.')

parser.add_argument('weight_paths', metavar='W', type=str, nargs='+', help='a weighted xlsx to compare')

if __name__ == "__main__":
    args = parser.parse_args()
    compare_weighted_regions(args.weight_paths[0], args.weight_paths[1])

    




