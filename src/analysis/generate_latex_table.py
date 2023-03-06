import argparse
import dataclasses
import math
import openpyxl
import collections

from typing import *


@dataclasses.dataclass
class ThresholdedRegionEntry:
    index: int
    ntuple: str
    weight: float
    rescaled_weight: float
    entropy: float
    average_entropy: float


def threshold_shannon(region_data_xlsx: str) -> Dict[int, List[ThresholdedRegionEntry]]:
    wb = openpyxl.load_workbook(region_data_xlsx, read_only=True)
    thresholded_region_data: dict[
        float, list[ThresholdedRegionEntry]
    ] = collections.OrderedDict()

    for i, ws_name in enumerate(wb.sheetnames):
        if i != 0:
            break

        count = 1
        entropy_sum = 0
        max_weight = 0

        for row in wb.worksheets[i].iter_rows(values_only=True):
            weight = float(row[len(row) - 1])

            if count == 1 or max_weight == 0:
                max_weight = weight

            rescaled_weight = weight / max_weight
            entropy = -rescaled_weight * math.log(rescaled_weight, 10)

            if weight in thresholded_region_data:
                entropy_sum += entropy
                avg_entropy = entropy_sum / count

                entry = ThresholdedRegionEntry(
                    count, row[0], weight, rescaled_weight, entropy, avg_entropy
                )
                thresholded_region_data[weight].append(entry)

                count += 1
                continue

            entropy_sum += entropy
            avg_entropy = entropy_sum / count

            tre = ThresholdedRegionEntry(
                count, row[0], weight, abs(rescaled_weight), abs(entropy), avg_entropy
            )
            thresholded_region_data[weight] = [tre]
            count += 1

    return thresholded_region_data


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate LaTeX table used in the paper."
    )
    parser.add_argument(
        "-r",
        "--region_data_xlsx",
        metavar="REGION_DATA",
        type=str,
        help="region data to be read",
        required=True,
    )

    args = parser.parse_args()

    thresholded_region_data = threshold_shannon(args.region_data_xlsx)

    print(
        f"""
\\begin{{tabular}}{{ { '|' + '|'.join([ "l" for i in range(6)]) + '|' } }}
\\hline
$4$-tuple & Weight & Rescaled Weight & Entropy & Average Entropy \\\\ \\hline
""",
        end="",
    )

    for k, v in thresholded_region_data.items():
        ntuples = [e.ntuple for e in v]
        print(v[0].index, end="")
        print("&", end="")
        print(", ".join(ntuples), end="")
        print("&", end="")
        print(f"{k:.5f}", end="")
        print("&", end="")
        print(f"{v[0].rescaled_weight:.5f}", end="")
        print("&", end="")
        print(f"{v[0].entropy:.5f}", end="")
        print("&", end="")
        print(f"{v[0].average_entropy:.5f}", end="")
        print("\\\\ \\hline")
    print("\\end{tabular}\n}")
