import openpyxl
import openpyxl.chart
import glob
import collections
import pathlib

def main() -> None:
	aggregate_folders = pathlib.Path('.').glob('aggregate_*')

	for folder in aggregate_folders:
		if not folder.is_dir():
			continue

		files = glob.glob(f'{folder}/*.xlsx')

		probabilities = collections.defaultdict(lambda : 0)

		for file in files:
			wb = openpyxl.load_workbook(file)
			ws = wb.active

			row_value_iter = iter(ws.values)
			next(row_value_iter) # Skip the column header

			for row in row_value_iter:
				base_position, probability = row
				base_position = int(base_position)
				probability = float(probability)

				probabilities[base_position] = probabilities[base_position] + probability

		number_of_files = len(files)

		for k in probabilities.keys():
			probabilities[k] /= number_of_files

		wb = openpyxl.Workbook()
		ws = wb.active

		ws['A1'] = "Base position"
		ws['B1'] = "Probability"

		for k in sorted(probabilities.keys()):
			ws[f'A{k + 1}'] = k
			ws[f'B{k + 1}'] = probabilities[k]


		chart = openpyxl.chart.LineChart()
		chart.title = "Average Probability vs Base Position"

		chart.x_axis_title = ws['A1']
		chart.y_axis_title = ws['B1']

		ref_probabilities = openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
		ref_base_positions = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

		chart.add_data(ref_probabilities)
		chart.set_categories(ref_base_positions)

		ws.add_chart(chart, "E5")

		wb.save(f'{folder}/{folder}_averages.xlsx')

if __name__ == "__main__":
	main()